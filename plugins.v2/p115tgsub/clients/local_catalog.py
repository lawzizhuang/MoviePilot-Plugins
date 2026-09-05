"""私有 XLSX 资源目录：只读加载、内存索引；不访问网络或持久化分享凭据。

依赖 openpyxl；配置容器内路径，由 SearchHandler 按订阅检索。
"""
import re
import unicodedata
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from app.log import logger


class LocalCatalog:
    REQUIRED = {"标题", "媒体标题", "链接", "访问码", "备注", "删除时间"}

    def __init__(self, path):
        self.path = Path(path)
        self._signature = None
        self._index = {}

    @staticmethod
    def _normalize(value):
        return re.sub(r"[\W_]", "", unicodedata.normalize("NFKC", str(value or "")).casefold())

    @staticmethod
    def _safe(value):
        text = re.sub(r"(?:https?://|magnet:|ed2k:)\S+", "", str(value or ""), flags=re.I)
        return re.sub(r"(?:访问码|提取码|密码|pwd|passcode)\s*[:：=]?\s*\w+", "", text, flags=re.I)

    @staticmethod
    def _share(link, password):
        try:
            parts = urlsplit(str(link or "").strip())
            if parts.scheme not in {"https", "http"} or parts.username or parts.password:
                return None
            host = (parts.hostname or "").lower()
            kind = "115" if host in {"115.com", "115cdn.com", "anxia.com"} else "quark" if host == "pan.quark.cn" else ""
            if not kind or not re.fullmatch(r"/s/[A-Za-z0-9]+/?", parts.path):
                return None
            query = dict(parse_qsl(parts.query))
            key = "password" if kind == "115" else "pwd"
            code = str(password or "").strip()
            if code:
                query[key] = code
            url = urlunsplit((parts.scheme, parts.netloc, parts.path, urlencode(query), parts.fragment))
            share_id = parts.path.strip("/").split("/")[-1]
            return kind, url, (kind, share_id, query.get(key, ""), parts.fragment)
        except (TypeError, ValueError):
            return None

    def _load(self):
        try:
            stat = self.path.stat()
            signature = (stat.st_mtime_ns, stat.st_size)
            if signature == self._signature:
                return
            from openpyxl import load_workbook
            book = load_workbook(self.path, read_only=True, data_only=False, keep_links=False)
            index, seen, count, deleted = {}, set(), 0, 0
            try:
                sheet = book["分享明细"]
                rows = sheet.iter_rows(values_only=True)
                headers = None
                for number, values in enumerate(rows, 1):
                    if headers is None:
                        if number > 20:
                            raise ValueError("missing headers")
                        if self.REQUIRED.issubset(set(values)):
                            headers = {value: i for i, value in enumerate(values) if value}
                        continue
                    def get(name):
                        pos = headers.get(name)
                        return values[pos] if pos is not None and pos < len(values) else None
                    if get("删除时间") not in (None, ""):
                        deleted += 1
                        continue
                    # 不计算公式，也不把公式字符串当作资源字段。
                    if any(isinstance(get(k), str) and get(k).startswith("=") for k in self.REQUIRED):
                        continue
                    share = self._share(get("链接"), get("访问码"))
                    if not share or share[2] in seen:
                        continue
                    title = self._safe(get("标题"))
                    media = self._safe(get("媒体标题"))
                    base = re.sub(r"\s*[（(](?:19|20)\d{2}[)）]\s*$", "", title)
                    names = {self._normalize(media), self._normalize(base)} - {""}
                    if not names:
                        continue
                    seen.add(share[2])
                    resource = {"url": share[1], "title": " ".join(dict.fromkeys(x for x in (title, media, self._safe(get("备注"))) if x)),
                                "text": "", "source": "local_catalog", "catalog_row": number,
                                "catalog_kind": share[0], "catalog_years": re.findall(r"[（(]((?:19|20)\d{2})[)）]", title)}
                    for name in names:
                        index.setdefault(name, []).append(resource)
                    count += 1
                if headers is None:
                    raise ValueError("missing headers")
            finally:
                book.close()
            after = self.path.stat()
            if signature != (after.st_mtime_ns, after.st_size):
                raise ValueError("file changed")
            self._index, self._signature = index, signature
            logger.info(f"[本地资源表] 加载完成：{count} 个115/夸克候选，排除删除记录 {deleted} 条")
        except Exception as exc:
            self._index, self._signature = {}, None
            # 异常正文可能包含单元格链接或密码，只显示异常类型。
            logger.warning(f"[本地资源表] 加载失败，本次不使用资源表：{type(exc).__name__}；请检查路径、权限、依赖及表头")

    def search(self, mediainfo, media_type, season=None, kind="115"):
        self._load()
        from app.schemas.types import MediaType
        year = str(getattr(mediainfo, "year", "") or "")
        output = []
        for resource in self._index.get(self._normalize(mediainfo.title), []):
            if resource["catalog_kind"] != kind:
                continue
            years = resource["catalog_years"]
            if year and years and year not in years:
                continue
            # 表格没有可靠的电影/电视剧类型列。电影要求明确年份；最终仍校验真实文件。
            if media_type == MediaType.MOVIE:
                if not year or year not in years:
                    continue
                if re.search(r"(?i)(?<![A-Za-z0-9])S\d{1,2}(?!\d)|第\s*\d+\s*[季集]|全\s*\d+\s*集", resource["title"]):
                    continue
            if media_type == MediaType.TV:
                seasons = {int(x) for x in re.findall(r"(?i)(?<![A-Za-z0-9])S(\d{1,2})(?!\d)", resource["title"])}
                if seasons and int(season or 1) not in seasons:
                    continue
            output.append(dict(resource))
            if len(output) >= 10:
                break
        if output:
            logger.info(f"[本地资源表] 命中 {len(output)} 个 {kind} 候选；来源行：{','.join(str(x['catalog_row']) for x in output)}；实际目录待验证")
        return output
